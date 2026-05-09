from __future__ import annotations

import io
from dataclasses import dataclass
from typing import Iterable

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.utils import column_index_from_string

from utils.excel_helpers import fill_formula_down, write_dataframe_to_sheet

AREAS = ["台北", "桃園", "新竹", "台中"]
TYPES = ["儲值金結算", "儲值金預收"]
LAST_COLS = {"儲值金結算": "T", "儲值金預收": "BJ"}

# 依照原本 Google Sheet「功能設定表」截圖與 Apps Script 邏輯整理。
# 這裡只先放儲值金結算 / 儲值金預收，其他工具以後可獨立新增。
DEFAULT_FORMULAS = [
    {"項目": "儲值金結算", "地區": "台北", "目標欄位": "V", "設定公式": "=YEAR(G2)"},
    {"項目": "儲值金結算", "地區": "桃園", "目標欄位": "V", "設定公式": "=YEAR(G2)"},
    {"項目": "儲值金結算", "地區": "新竹", "目標欄位": "V", "設定公式": "=YEAR(G2)"},
    {"項目": "儲值金結算", "地區": "台中", "目標欄位": "V", "設定公式": "=YEAR(G2)"},
    {"項目": "儲值金結算", "地區": "台北", "目標欄位": "W", "設定公式": "=W$1-Q2"},
    {"項目": "儲值金結算", "地區": "桃園", "目標欄位": "W", "設定公式": "=W$1-Q2"},
    {"項目": "儲值金結算", "地區": "新竹", "目標欄位": "W", "設定公式": "=W$1-Q2"},
    {"項目": "儲值金結算", "地區": "台中", "目標欄位": "W", "設定公式": "=W$1-Q2"},
    {"項目": "儲值金結算", "地區": "台北", "目標欄位": "X", "設定公式": '=IF(OR(B2="檸檬清潔代客保留",B2="Wayne",B2="檸檬保留單"),"",N2)'},
    {"項目": "儲值金結算", "地區": "桃園", "目標欄位": "X", "設定公式": '=IF(OR(B2="代客保留",B2="Wayne"),"",N2)'},
    {"項目": "儲值金結算", "地區": "新竹", "目標欄位": "X", "設定公式": '=IF(OR(B2="新竹代客保留",B2="Wayne"),"",N2)'},
    {"項目": "儲值金結算", "地區": "台中", "目標欄位": "X", "設定公式": '=IF(OR(B2="檸檬保留",B2="檸檬台中保留"),"",N2)'},
    {"項目": "儲值金結算", "地區": "台北", "目標欄位": "AM", "設定公式": "=B2&E2"},
    {"項目": "儲值金結算", "地區": "桃園", "目標欄位": "AM", "設定公式": "=B2&E2"},
    {"項目": "儲值金結算", "地區": "新竹", "目標欄位": "AM", "設定公式": "=B2&E2"},
    {"項目": "儲值金結算", "地區": "台中", "目標欄位": "AM", "設定公式": "=B2&E2"},
    {"項目": "儲值金結算", "地區": "台北", "目標欄位": "AN", "設定公式": '=IF(OR(B2="檸檬清潔代客保留",B2="Wayne",B2="檸檬保留單"),"",N2)'},
    {"項目": "儲值金結算", "地區": "桃園", "目標欄位": "AN", "設定公式": '=IF(OR(B2="代客保留",B2="Wayne"),"",N2)'},
    {"項目": "儲值金結算", "地區": "新竹", "目標欄位": "AN", "設定公式": '=IF(OR(B2="新竹代客保留",B2="Wayne"),"",N2)'},
    {"項目": "儲值金結算", "地區": "台中", "目標欄位": "AN", "設定公式": '=IF(OR(B2="檸檬保留",B2="檸檬台中保留"),"",N2)'},
    {"項目": "儲值金結算", "地區": "台北", "目標欄位": "AR", "設定公式": "=IFNA(VLOOKUP(AP2,'台北儲值金預收'!BO:BP,2,FALSE),0)"},
    {"項目": "儲值金結算", "地區": "桃園", "目標欄位": "AR", "設定公式": "=IFNA(VLOOKUP(AP2,'桃園儲值金預收'!BO:BP,2,FALSE),0)"},
    {"項目": "儲值金結算", "地區": "新竹", "目標欄位": "AR", "設定公式": "=IFNA(VLOOKUP(AP2,'新竹儲值金預收'!BO:BP,2,FALSE),0)"},
    {"項目": "儲值金結算", "地區": "台中", "目標欄位": "AR", "設定公式": "=IFNA(VLOOKUP(AP2,'台中儲值金預收'!BO:BP,2,FALSE),0)"},
    {"項目": "儲值金結算", "地區": "台北", "目標欄位": "AS", "設定公式": "=AQ2+AR2"},
    {"項目": "儲值金結算", "地區": "桃園", "目標欄位": "AS", "設定公式": "=AQ2+AR2"},
    {"項目": "儲值金結算", "地區": "新竹", "目標欄位": "AS", "設定公式": "=AQ2+AR2"},
    {"項目": "儲值金結算", "地區": "台中", "目標欄位": "AS", "設定公式": "=AQ2+AR2"},
    {"項目": "儲值金預收", "地區": "台北", "目標欄位": "BL", "設定公式": "=M2&N2"},
    {"項目": "儲值金預收", "地區": "桃園", "目標欄位": "BL", "設定公式": "=M2&N2"},
    {"項目": "儲值金預收", "地區": "新竹", "目標欄位": "BL", "設定公式": "=M2&N2"},
    {"項目": "儲值金預收", "地區": "台中", "目標欄位": "BL", "設定公式": "=M2&N2"},
    {"項目": "儲值金預收", "地區": "台北", "目標欄位": "BM", "設定公式": "=AA2+AB2"},
    {"項目": "儲值金預收", "地區": "桃園", "目標欄位": "BM", "設定公式": "=AA2+AB2"},
    {"項目": "儲值金預收", "地區": "新竹", "目標欄位": "BM", "設定公式": "=AA2+AB2"},
    {"項目": "儲值金預收", "地區": "台中", "目標欄位": "BM", "設定公式": "=AA2+AB2"},
    {"項目": "儲值金預收", "地區": "台北", "目標欄位": "BQ", "設定公式": "=VLOOKUP(BO2,'台北儲值金結算'!AP:AP,1,FALSE)"},
    {"項目": "儲值金預收", "地區": "桃園", "目標欄位": "BQ", "設定公式": "=VLOOKUP(BO2,'桃園儲值金結算'!AP:AP,1,FALSE)"},
    {"項目": "儲值金預收", "地區": "新竹", "目標欄位": "BQ", "設定公式": "=VLOOKUP(BO2,'新竹儲值金結算'!AP:AP,1,FALSE)"},
    {"項目": "儲值金預收", "地區": "台中", "目標欄位": "BQ", "設定公式": "=VLOOKUP(BO2,'台中儲值金結算'!AP:AP,1,FALSE)"},
]


@dataclass
class SourceFile:
    area: str
    kind: str
    filename: str
    df: pd.DataFrame


def detect_area_type(filename: str) -> tuple[str | None, str | None]:
    area = next((a for a in AREAS if a in filename), None)
    kind = next((t for t in TYPES if t in filename), None)
    return area, kind


def read_uploaded_file(uploaded_file) -> pd.DataFrame:
    name = uploaded_file.name.lower()
    if name.endswith(".csv"):
        return pd.read_csv(uploaded_file, header=None)
    if name.endswith(".xls"):
        return pd.read_excel(uploaded_file, header=None, engine="xlrd")
    return pd.read_excel(uploaded_file, header=None, engine="openpyxl")


def remove_blank_rows(df: pd.DataFrame) -> pd.DataFrame:
    return df.dropna(how="all").reset_index(drop=True)


def trim_to_source_range(df: pd.DataFrame, kind: str) -> pd.DataFrame:
    last_col_number = column_index_from_string(LAST_COLS[kind])
    df = remove_blank_rows(df)
    return df.iloc[:, :last_col_number].copy()


def apply_formulas(ws, formula_rows: pd.DataFrame, data_row_count: int) -> None:
    for _, row in formula_rows.iterrows():
        col = str(row["目標欄位"]).strip().upper()
        formula = str(row["設定公式"]).strip()
        if not col or not formula or formula.lower() == "nan":
            continue
        ws.cell(1, column_index_from_string(col), col)
        fill_formula_down(ws, col, formula, data_row_count)


def clear_bp_bq_when_bo_is_total(ws, data_row_count: int) -> None:
    bo_col = column_index_from_string("BO")
    bp_col = column_index_from_string("BP")
    bq_col = column_index_from_string("BQ")
    for row_idx in range(2, data_row_count + 2):
        if ws.cell(row_idx, bo_col).value == "總和":
            ws.cell(row_idx, bp_col, None)
            ws.cell(row_idx, bq_col, None)


def write_settings_sheet(wb: Workbook, formulas: pd.DataFrame) -> None:
    ws = wb.create_sheet("功能設定表")
    headers = list(formulas.columns)
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(1, col_idx, header)
    for row_idx, values in enumerate(formulas.itertuples(index=False, name=None), start=2):
        for col_idx, value in enumerate(values, start=1):
            ws.cell(row_idx, col_idx, value)


def create_workbook(source_files: Iterable[SourceFile], formulas: pd.DataFrame) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)
    write_settings_sheet(wb, formulas)

    sheets = {}
    for area in AREAS:
        for kind in TYPES:
            sheet_name = f"{area}{kind}"
            sheets[sheet_name] = wb.create_sheet(sheet_name)

    report_rows = []
    for src in source_files:
        sheet_name = f"{src.area}{src.kind}"
        ws = sheets[sheet_name]
        df = trim_to_source_range(src.df, src.kind)
        write_dataframe_to_sheet(ws, df, start_row=2, start_col=1)

        matched = formulas[(formulas["項目"] == src.kind) & (formulas["地區"] == src.area)]
        apply_formulas(ws, matched, len(df))

        if src.kind == "儲值金預收":
            clear_bp_bq_when_bo_is_total(ws, len(df))

        report_rows.append([src.filename, src.area, src.kind, len(df), len(matched), "完成"])

    report = wb.create_sheet("診斷報表")
    headers = ["檔名", "地區", "項目", "資料列數", "套用公式數", "狀態"]
    for col_idx, header in enumerate(headers, start=1):
        report.cell(1, col_idx, header)
    for row_idx, values in enumerate(report_rows, start=2):
        for col_idx, value in enumerate(values, start=1):
            report.cell(row_idx, col_idx, value)

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def render() -> None:
    st.title("儲值金管理")
    st.caption("處理儲值金結算與儲值金預收：上傳各區原始檔 → 套用功能設定表公式 → 下載 Excel。")

    with st.sidebar:
        st.divider()
        st.subheader("儲值金管理說明")
        st.markdown(
            """
            1. 檔名需包含地區：台北 / 桃園 / 新竹 / 台中  
            2. 檔名需包含類型：儲值金結算 / 儲值金預收  
            3. 可在畫面上直接調整公式  
            4. 下載後用 Excel 或 Google 試算表開啟
            """
        )

    uploaded_files = st.file_uploader(
        "上傳 xlsx / xls / csv，可一次多選",
        type=["xlsx", "xls", "csv"],
        accept_multiple_files=True,
    )

    formula_df = pd.DataFrame(DEFAULT_FORMULAS)
    with st.expander("功能設定表（可修改）", expanded=True):
        st.info("公式請以第 2 列為基準，例如 =YEAR(G2)、=AQ2+AR2。")
        formula_df = st.data_editor(
            formula_df,
            num_rows="dynamic",
            use_container_width=True,
            column_config={
                "項目": st.column_config.SelectboxColumn(options=TYPES),
                "地區": st.column_config.SelectboxColumn(options=AREAS),
            },
        )

    source_files: list[SourceFile] = []
    skipped = []
    if uploaded_files:
        for file in uploaded_files:
            area, kind = detect_area_type(file.name)
            if not area or not kind:
                skipped.append([file.name, "檔名需包含地區與儲值金結算/儲值金預收"])
                continue
            try:
                source_files.append(SourceFile(area, kind, file.name, read_uploaded_file(file)))
            except Exception as exc:
                skipped.append([file.name, str(exc)])

    if source_files:
        st.subheader("已辨識檔案")
        st.dataframe(
            pd.DataFrame(
                [
                    {"檔名": item.filename, "地區": item.area, "項目": item.kind, "讀入列數": len(item.df)}
                    for item in source_files
                ]
            ),
            use_container_width=True,
        )

    if skipped:
        st.warning("部分檔案略過，請檢查命名或格式。")
        st.dataframe(pd.DataFrame(skipped, columns=["檔名", "原因"]), use_container_width=True)

    if st.button("產生儲值金管理 Excel", type="primary", disabled=not bool(source_files)):
        output = create_workbook(source_files, formula_df)
        st.success("已產生 Excel。")
        st.download_button(
            "下載 儲值金管理.xlsx",
            data=output,
            file_name="儲值金管理_output.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
